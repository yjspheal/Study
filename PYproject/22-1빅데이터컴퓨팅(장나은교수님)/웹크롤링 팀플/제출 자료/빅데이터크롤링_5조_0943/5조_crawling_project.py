#!/usr/bin/env python
# coding: utf-8

#####20171228 박현진 #########################################################################################################
#oliveyoung_crawling.py
import urllib.request
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from PIL import Image
import re
from bs4 import BeautifulSoup
from selenium import webdriver
from tqdm import tqdm
from selenium.webdriver.common.keys import Keys
keys=Keys()
import time
import os
import warnings
warnings.filterwarnings(action='ignore')

###파이프라인 1. 올리브영 랭킹 페이지 제품 100개 리스트 만들기 
def crawling_top100_list():
    #판매랭킹 베이스 url 
    main_url = 'https://www.oliveyoung.co.kr/store/main/getBestList.do'

    #크롬 드라이버 열기 
    driver = webdriver.Chrome(".\chromedriver.exe") 
    driver.get(main_url)

    #로딩 될때까지 기다리기 
    driver.implicitly_wait(5) # seconds

    #'스킨케어' 클릭 
    skincare_button='#Container > div.best-area > div.common-menu > ul > li:nth-child(2) > button'
    key = driver.find_element_by_css_selector(skincare_button)
    key.click()
    time.sleep(3)

    #상품 상세정보를 저장할 데이터프레임 
    df=pd.DataFrame(columns=['순위','상품이름','브랜드','가격','상세페이지'])

    #각 라인마다 4개의 원소를 들고올것, 반복횟수: (ul:nth-child(1)~ul:nth-child(25))
    rank=1
    for line in range(1, 26):

        prod_box='#Container > div.best-area > div.TabsConts.on > ul:nth-child('+str(line)+')'

        for idx in range(1,5):
            if idx %4==1: ##첫번째 박스만 태그가 'flag'로 되어 있으므로 예외처리 
                nth_box= ' > li.flag'

            elif idx %4==0:#네번째박스
                nth_box = '> li:nth-child('+str(4)+')'

            else : ##두,세번째 박스 
                nth_box = '> li:nth-child('+str(idx%4)+')'
            #상세페이지 url
            prod_detail_tag = prod_box + nth_box+ ' > div > a'
            #브랜드
            brand_tag=  prod_box + nth_box+  ' > div > div > a > span'
            #상품이름
            prod_name_tag=  prod_box + nth_box+  ' div > div > a > p'
            #가격
            price_tag =  prod_box + nth_box+  ' > div > p.prd_price > span.tx_cur > span'
            #상품이미지
            img_tag= prod_box + nth_box +' > div > a > img'
            
            #태그로 읽어들인 텍스트에서 속성추출, 전처리 
            prod_detail_url = driver.find_element_by_css_selector(prod_detail_tag).get_attribute('href')
            brand = driver.find_element_by_css_selector(brand_tag).text
            prod_name = driver.find_element_by_css_selector(prod_name_tag).text
            price =  int(driver.find_element_by_css_selector(price_tag).text.replace(',',''))
            df= df.append(pd.DataFrame([[rank, prod_name, brand, price, prod_detail_url]],columns=['순위','상품이름','브랜드','가격','상세페이지']), ignore_index=True)
            
            #크롤링 확인을 위한 코드 
            print(brand+':'+prod_name)
            print(prod_detail_url)

            #크롤링한 이미지 저장
            img = driver.find_element_by_css_selector(img_tag).get_attribute('src')
            #1. 이미지 저장 폴더 설정 (없으면 폴더 생성)
            try:
                os.mkdir("new_crawled_img") 
            except FileExistsError:
                print('new_crawled_img 파일이 이미 존재합니다. 해당 폴더에 이미지를 크롤링합니다.')
            filename = 'new_crawled_img/'+str(rank)+'_'+brand+'.jpg'

            #2. 경로내 중복되는 파일 명 확인
            try:
                # 중복되는 파일 명이 없다면 패스
                if not os.path.isfile(filename):
                    pass
                # 중복된다면 문구 출력 후 다음 이미지로 넘어감
                else:
                    print('이전에 다운로드한 이미지가 존재합니다.')
                    pass
            except OSError:
                print ('os error')
                #sys.exit(0)

            #3. 이미지 저장
            try:
                req = urllib.request.Request(img, headers={'User-Agent': 'Mozilla/5.0'})
                imgUrl = urllib.request.urlopen(req).read() #웹 페이지 상의 이미지를 불러옴
                with open(filename,"wb") as f: #디렉토리 오픈
                    f.write(imgUrl) #파일 저장

            except urllib.error.HTTPError:
                print('이미지 저장 에러')
                pass 

            #반복문에서 순위를 계산중이므로, 하나의 상품에 대한 크롤링이 전부 끝나면 순위 올려주기 
            rank+=1

            time.sleep(1)
    df.to_excel('(new)올리브영랭크100위화장품정보.xlsx')


##파이프라인 2. 100개 리스트 행마다 반복하며 상세페이지 들어가서 상세 정보 크롤링 
def iter_dataframe(df):
    def crawl_olive(main_url):
        driver = webdriver.Chrome(".\chromedriver.exe") 
        time.sleep(3)
        driver.get(main_url)
        driver.implicitly_wait(5) 
        key = driver.find_element_by_css_selector('#reviewInfo > a')

        key.click()
        time.sleep(3)
        
        try: #상세리뷰가 있거나, 리뷰의 개수가 있는 경우에만 긁어야 에러가 나지 않으므로 try안에 넣음
            rating = driver.find_element_by_css_selector('#gdasContentsArea > div > div.product_rating_area > div > div.star_area')
            grpah = driver.find_element_by_css_selector('#gdasContentsArea > div > div.product_rating_area > div > div.graph_area')
            detail = driver.find_element_by_css_selector('#gdasContentsArea > div > div.poll_all.clrfix')

            rating_text= rating.text.split('\n')
            graph_text= grpah.text.split('\n')
            detail_text= detail.text.split('\n')
            review_n = int(re.sub(r'[^0-9]', '', rating_text[0]))
            star = float(rating_text[1][:3])
        except:#상세리뷰가 없거나 리뷰가 없는 상품 -> 결측치로 이후에 처리하기 위해 -1값을 대신 넣음 
            review_n=-1
            star=-1
            graph_text=""
            detail_text = ""
            print('상세리뷰가 없거나 리뷰가 없는 상품입니다.:',main_url)
            pass
        return review_n, star, graph_text, detail_text
    
    # 빈 DataFrame 생성하기
    df2=pd.DataFrame(columns=['순위','리뷰개수','별점','별점그래프','타입그래프'])
    
    #받은 올리브영 상품 리스트 데이터프레임에 대해서 반복하면서 상세 리뷰를 채워주는 코드 
    for i in tqdm(df.index):
        rank=i+1
        review_n, star, graph_text, detail_text= crawl_olive(df['상세페이지'][i])
        df2 = df2.append(pd.DataFrame([[rank,review_n, star, graph_text, detail_text]], columns=['순위','리뷰개수','별점','별점그래프','타입그래프']), ignore_index=True)
    
    #원래의 올리브영 데이터프레임과 새로 크롤링한 상세 리뷰 데이터를 합집합으로 머지해주는 코드 
    merge_df= pd.merge(df, df2, on='순위',how='outer')
    merge_df.drop(['Unnamed: 0'], axis = 1, inplace = True)

    merge_df.to_excel('new_올리브영랭크100위화장품정보_상세페이지.xlsx')
    
def olive_main():
    print('1차 크롤링 : 올리브영랭크 상위 100위화장품을 크롤링합니다.')
    crawling_top100_list()
    df=pd.read_excel('원본_올리브영랭크100위화장품정보.xlsx')
    print('2차 크롤링 : 올리브영랭크100위화장품정보를 읽어 상세페이지를 크롤링합니다.')
    result_df= iter_dataframe(df)
    print('올리브영 크롤링이 완료되었습니다.')


#####20180488 김민서 #########################################################################################################
from urllib.request import *
from bs4 import BeautifulSoup
import ssl
from selenium import webdriver
from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import pandas as pd
import time

ssl._create_default_https_context = ssl._create_unverified_context

olive = pd.read_excel(r".\원본_올리브영랭크100위화장품정보_상세페이지.xlsx")

#브랜드 이름 추출하는 함수
def product_name(file):
  name_list = [] #브랜드 이름을 담을 리스트
  for i in range(len(file)):
    product_name = file['브랜드'][i] #리스트에 브랜드 저장
    name_list.append(product_name) 
    name_list = list(set(name_list))

  return name_list


#브랜드에 대한 정보들을 알아보는 함수(트렌드 함수)
def trend_index(product): #각 브랜드 이름 인자로 받기
    
    #첫페이지 열기
    driver = webdriver.Chrome('.\chromedriver.exe')
    driver.implicitly_wait(10) # seconds
    driver.get("https://www.itemscout.io/keyword")
    driver.implicitly_wait(10) # seconds
    time.sleep(2) #로딩이 느릴때 쓰기
    
    #키워드 입력, Enter
    keyword = driver.find_element_by_class_name("input-keyword")
    keyword.send_keys(product) #브랜드 이름 입력해서 엔터

    enter = '//*[@id="app"]/div/main/div/div/div/div[1]/div[1]/div[1]/div[2]/div[1]/div'
    enter_click = driver.find_element_by_xpath(enter).click()

    driver.implicitly_wait(10) # seconds
    time.sleep(0.5)

    #한 달 검색수
    pre_elem = driver.find_elements_by_class_name("count-stat")[1]
    search_n = pre_elem.text
    search_n = search_n.replace(',','')

    #Top40 6개월 매출
    pre_elem = driver.find_elements_by_class_name("count-stat")[2]
    sales = pre_elem.text
    sales = sales.replace(",","")
    
    #Top40 6개월 판매량
    pre_elem = driver.find_elements_by_class_name("count-stat")[3]
    volume = pre_elem.text
    volume = volume.replace(",","")

    #Top40 평균 가격
    pre_elem = driver.find_elements_by_class_name("count-stat")[4]
    price = pre_elem.text
    price = price.replace(",","")
    
    driver.close() #다 추출하고 나면 창 닫기
    
    return search_n, sales, volume, price

def process(trend): #전처리 함수
    
    search_n = []
    for i in trend['한 달 검색수']:
        i = int(i.replace("회",""))
        search_n.append(i)
        
    sales = []
    for i in trend['6개월 매출']:
        if i[:-2] == '만원' : #만원단위는 0네개 붙이고 int
            i = i[:-2]
            i=i+'0000'
        else:
            i = i[:-2]
            i=i+'00000000' #억 단위는 0여덟개 붙이고 int
        sales.append(int(i))
        
    sales_volume = []
    for i in trend['6개월 판매량']:
        i = int(i.replace("개",""))
        sales_volume.append(i)
        
    price = []
    for i in trend['평균 가격']:
        i = int(i.replace("원",""))
        price.append(i)
        
    trend['한 달 검색수'] = search_n
    trend['6개월 매출'] = sales
    trend['6개월 판매량'] = sales_volume
    trend['평균 가격'] = price
    
    

def item_crawl_main():
    n_list = product_name(olive) #olive파일의 브랜드 이름이 담긴 리스트
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Trend Index"
    ws1.append(["브랜드", "한 달 검색수", "6개월 매출", "6개월 판매량", "평균 가격"])

    for i in n_list: #겹치지 않는 브랜드 명을 담은 리스트를 인자로 받음
        search_n, sales, volume, price = trend_index(i) #함수 호출
        ws1.append([i, search_n, sales, volume, price]) #정보 저장

    wb.save(filename='new_트렌드지수 크롤링(브랜드).xlsx') #끝나고 나면 엑셀로 만들기
    
    trend = pd.read_excel(r".\new_트렌드지수 크롤링(브랜드).xlsx")
    process(trend) #전처리하여 사용할 수 있게 데이터 준비

#####20191245 노유정 #########################################################################################################

import time
import re
import datetime as dt
import pandas as pd
import openpyxl as xl
import os
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from tqdm import tqdm

def get_youtube_data(brand, goods):
    search_goods = f'{brand} {goods}'
    
    # 드라이버 주소 설정
    browser = webdriver.Chrome(".\chromedriver.exe") 
    browser.implicitly_wait(1)   
    
    # 크롤링 데이터를 저장할 경로 및 엑셀 파일 이름 지정
    try:      # 데이터 100개를 저장할 폴더 생성('YT_data') 및 이미 존재하는 폴더일 시 예외처리
        os.mkdir("new_YT_data") 
    except FileExistsError:
        print('YT_data 폴더가 이미 존재합니다. 해당 폴더에 데이터를 저장합니다.')
    result = pd.ExcelWriter("./new_YT_data/new_" + search_goods + '.xlsx', engine='openpyxl')
    
    # 유튜브 검색을 위한 url을 만들어 열고 body 추출
    yt_url = "https://www.youtube.com"
    search_goods = search_goods.replace(' ','+')  # url용으로 제품명 사이 공백 +로 대체
    target_url  = yt_url + "/results?search_query="+search_goods + "+리뷰" + "&sp=CAI%253D"      # 업로드날짜 필터링
    browser.get(target_url)
    body = browser.find_element_by_tag_name('body')
    
    # 전체 데이터를 추출하기 위하여 페에지를 쭉 내림. 필터링이 되기 때문에 50번이면 충분
    for pg_down in range(50):  # 페이지 다운 수
        body.send_keys(Keys.PAGE_DOWN)
        browser.implicitly_wait(1)
    
    # 해당 페이지 html 소스를 beautifulsoup을 이용하여 html 에 저장
    html0 = browser.page_source
    html = BeautifulSoup(html0,'html.parser')
    
    # 검색 결과창에서 각 비디오 하나씩을 나누어줌
    video_datas = html.find_all('ytd-video-renderer',{'class':'style-scope ytd-item-section-renderer'})

    
    #데이터 저장을 위한 기본 데이터프레임 생성
    dataframe = pd.DataFrame({'title':[], 'youtube_url':[], 'subscribers':[], 'views':[], "nice":[], "reply":[]})
    
    # 비디오 url을 저장할 리스트 생성(해당 url은 youtube.com/ 이후에 들어가는 부분)
    video_url_list = []
    for i in range(len(video_datas)):
        # 검색 결과 창에서 얻을 수 있는 데이터인 동영상 제목(title)과 동영상 링크 추출
        title = video_datas[i].find('a',{'id':'video-title'}).get_text() 
        url = yt_url + video_datas[i].find('a',{'id':'thumbnail'})['href'] 
        video_url_list.append(url)
        
        # 상세 동영상 페이지 접속 및 바디 추출
        cur_url = video_url_list[i]
        browser.get(cur_url); time.sleep(5)
        body = browser.find_element_by_tag_name('body')
        
        # 댓글 수를 보려면 페이지를 동영상 목록까지 내린 후 기다려야 나오므로 페이지를 내리고 기다리는 과정을 반복
        for pg_down in range(10):
            body.send_keys(Keys.PAGE_DOWN)
            time.sleep(0.5)
        
        # 현재 동영상 페이지 소스코드를 beautifulsoup 이용하여 html에 저장
        html0 = browser.page_source
        html = BeautifulSoup(html0,'html.parser')
        
        # 조회수, 좋아요수, 댓글수는 영상마다 감춰두거나 막아놓는 경우가 있어 에러가 발생하여 예외처리
        try:
            
            # 조회수를 크롤링하고, nnn,nnn회 라는 결과를 얻게 되어 replace로 정리하여 정수로 저장
            view_count = html.find('span',{'class':'view-count style-scope ytd-video-view-count-renderer'}).get_text().split()[1]
            view_count = view_count.replace(",", ""); view_count = view_count.replace("회", ""); view_count = int(view_count)
            
            # 채널명(유튜버)를 크롤링하여 youtuber에 저장
            channel = html.find("div", {'class' : 'style-scope ytd-channel-name'})     # 채널명을 나타내는 태그와 class는 여러 곳에서 쓰이므로, class style-scope ytd-channel-name를 이용하여 chammel에 저장 후 chammel에서 해당 형태를 다시 찾아냄
            youtuber = channel.find("a", {"class" : 'yt-simple-endpoint style-scope yt-formatted-string'}).get_text()
            
            # 데이터 정제를 위하여 조회수가 500 미만이거나 채널 명에 브랜드가 포함되는 경우(즉 브랜드 자체 광고 동영상인 경우) 해당 동영상은 넘기고 다음 '동영상'으로 넘어감
            if view_count < 500 or youtuber.find(brand) != -1:
                continue       
            
            # 업로드 일자를 크롤링하여 만약 현재(코드를 돌리는 시점)부터 6개월 이전에 만들어진 영상이라면 break하고 다음 '제품'으로 넘어감(업로드 일자를 기준으로 정렬된 동영상들이기 때문에 이 이후는 볼 필요가 없으므로 break)
            upload_date = html.find_all("yt-formatted-string", {"class" : "style-scope ytd-video-primary-info-renderer"})[1].get_text()
            daybefore = dt.datetime.today() - dt.timedelta(days = 180) 
            upload_date = upload_date[:len(upload_date)-1]
            ul_year, ul_month, ul_day = map(int,upload_date.split("."))
            if daybefore > dt.datetime(ul_year, ul_month, ul_day):   # 업로드 6개월 이전이라면 break
                break
            
            # 좋아요 수 크롤링, 좋아요 수는 싫어요 수와 같은 태그 및 class를 가지고 그중 첫번째에 해당함
            nice = html.find_all("yt-formatted-string",{"class" : "style-scope ytd-toggle-button-renderer style-text"})[0]
            # get_text를 하면 'n.n만개'처럼 축약된 형태가 반환되고, 태그 속 aria-label에 상세한 좋아요 수가 들어있어 이를 반환
            nice_count = nice["aria-label"].split()[-1]     # '좋아요 nnnnn개'의 형태가 반환되므로 split함
        
            # 좋아요 수가 100개 이하인 경우 콤마가 포함되지 않는 숫자가 반환되어 replace가 에러날 수 있으므로 문자로 변환하였다가 정제 후 정수로 다시 변환하여 저장
            nice_count = str(nice_count)
            nice_count = nice_count.replace(",", ""); nice_count = nice_count.replace("개", "")
            nice_count = int(nice_count)
            
            # 구독자 수를 아래의 태그와 id를 이용하여 크롤링
            subscribers = html.find("yt-formatted-string",{"id" : "owner-sub-count"}).get_text().split()[1]
            
            # 댓글 수 크롤링. 댓글 수를 나타내는 태그와 class는 아래의 답글 수 등에서도 계속 사용되기 때문에 info에 댓글 수가 있는 comments-header를 추출하여 저장 후 사용
            info = html.find('h2', {'class' : 'style-scope ytd-comments-header-renderer'})
            reply_count = info.find_all('span', {'class' : 'style-scope yt-formatted-string'})[1].get_text()    #info에는 댓글/nnn/개 의 소스가 저장되어있어 2번째 인자 추출하여 get_text
            reply_count = int(reply_count)    # nnn의 형태이나 문자이므로 정수로 변환하여 저장
            
            # 각 데이터를 합쳐 데이터프레임을 만들고 insert_data에 저장
            insert_data = pd.DataFrame({'title':[title], 'youtube_url':[url], 'subscribers':[subscribers], 'views':[view_count], "nice":[nice_count], "reply":[reply_count]})
            print("추출 Yes\n")  # 코드 실행 중 가시화를 위한 코드
            
        except Exception as e:
            continue 
         
        # insert_data를 위에 만들었던 dataframe에 합쳐넣고, 이를 엑셀로 변환하여 저장
        dataframe = dataframe.append(insert_data)
        dataframe.to_excel(result, index = False)
        result.save()
    
    # dataframe에 들어가는 데이터가 없는 경우, 즉 최근 6개월 이내에 올라온 조회수 500 이상의 리뷰 동영상이 없는 경우 엑셀이 제대로 저장되지 않으므로 데이터프레임을 하나 만들어서 저장
    if dataframe.shape[0] == 0:
        dataframe = pd.DataFrame({'title': [0], 'youtube_url': [0], 'subscribers':[0], 'views':[0], "nice":[0], "reply":[0]})   # 조회수가 0인 영상을 존재하지 않기 때문에 시각화 과정에서 이를 이용하여 실제 데이터와 분리
        dataframe.to_excel(result, index = False)
        result.save()
    

def youtube_crawling():
    # 제품 명과 브랜드 이름이 들어있는 엑셀 파일을 불러와 활성화
    raw_data = xl.load_workbook('원본_올리브영랭크100위화장품정보_상세페이지.xlsx')
    raw_data = raw_data.active
    
    # C, D열에서 각각 제품 명과 브랜드 이름(이 들어있는 셀 리스트)를 slice하여 저장
    goods_name = raw_data["D"][1::]
    brand_name = raw_data["E"][1::]

    # get_youtube_data 함수에 제품명과 브랜드를 넣어 실행
    for i in range(1,100):
        print(f'rank {i+1}',end = ' ')  # 코드 실행 중 가시화를 위한 코드
        goods = goods_name[i].value   # _name[i] 자체는 셀이기 때문에 value로 셀의 내용을 추출하여 저장
        brand = brand_name[i].value
        get_youtube_data(brand, goods)
        
        
# 기존 엑셀에 유튜브 데이터 추가하는 함수
def excel_add_youtube():
    # 기존 정보 들어있는 엑셀을 load하고 활성화시킴
    raw_datad = xl.load_workbook('원본_올리브영랭크100위화장품정보_상세페이지.xlsx')
    raw_data = raw_datad.active
    
    # 저장된 100개의 파일 불러오기 위한 브랜드, 제품명 가져오기
    goods_name = raw_data["C"][1::]; brand_name = raw_data["D"][1::]
    
    # format을 맞추기 위해 첫번째 행에 들어갈 값들 지정
    raw_data["Q1"] = "유튜브 컨텐츠 수"; raw_data["R1"] = "총 조회수"; raw_data["S1"] = "총 좋아요수"; raw_data["T1"] = "총 댓글수"
    
    # 100개의 파일을 불러와 기존 엑셀에 대하여 유튜브 컨텐츠 수, 총 조회수, 좋아요수, 댓글수 추출하여 저장
    for i in range(100):
        goods = goods_name[i].value
        brand = brand_name[i].value

        filed = xl.load_workbook('./원본_YT_data/'+f'{brand} {goods}'+".xlsx")  # 파일 이름으로 불러오므로 폴더 내에 데이터 순서에 무관하게 각 파일을 가져옴
        file = filed.active
        
        # 각 파일 내에서 조회수, 좋아요수, 댓글 수 들어있는 열을 가져옴
        views = file["D"][1::]; nices = file["E"][1::]; replies = file["F"][1::] 
        
        # 컨텐츠의 갯수는 파일의 행 갯수로 가져옴
        contents_count = len(file["E"][1::])
        
        # 각 셀의 value를 가져와 총 합계를 각각 구함
        views_sum = 0; nices_sum = 0; replies_sum = 0
        for v in views:
            views_sum += v.value
        
        for n in nices:
            nices_sum += n.value
        
        for r in replies:
            replies_sum += r.value

    # 구한 값들을 차례로 엑셀 칸에 집어넣음
    raw_data[f"Q{i+2}"] = contents_count; raw_data[f"R{i+2}"] = views_sum; raw_data[f"S{i+2}"] = nices_sum; raw_data[f"T{i+2}"] = replies_sum
    
    # 만든 엑셀 데이터를 새로운 이름으로 저장
    raw_datad.save('new_올리브영랭크100위화장품정보+유튜브데이터.xlsx')
    
def make_top20():
    # 브랜드 top 20 리스트 생성(제품 top 20으로 브랜드 갯수는 더 적음)
    tops = ['아누아','토리든','닥터지','라로슈포제','라운드랩','아벤느','더랩바이블랑두','디오디너리','셀리맥스','브링그린','에스트라']
    
    try: # 결과 저장할 엑셀 생성
        top20d = pd.ExcelWriter('new_올리브영_진정_상위20.xlsx', engine='openpyxl')
        dataframe = pd.DataFrame(['','',' ','','','','','',''])
        dataframe.to_excel(top20d, index = False)
        top20d.save()
    except:
        print("이미 파일이 존재합니다. 기존 파일에 덮어쓰기를 시작합니다.")
    
    # 만든 파일 활성화
    top20d = xl.load_workbook('new_올리브영_진정_상위20.xlsx')
    top20 = top20d.active
    
    # 유튜브데이터까지 들어있는 엑셀을 load후 활성화
    ol_yud = xl.load_workbook('원본_올리브영랭크100위화장품정보+유튜브데이터.xlsx')
    ol_yu = ol_yud.active
    
    # 트렌드 지수 들어있는 엑셀 load 후 활성화
    trendd = xl.load_workbook('./원본_트렌드지수 크롤링(브랜드).xlsx')
    trend = trendd.active

    # format을 맞추기 위해 첫번째 행에 들어갈 값들 지정
    top20["B1"] = "브랜드"; top20["C1"] = "평균가격"; top20["D1"] = "평균리뷰개수"; top20["E1"] = "평균별점"
    top20["F1"] = "평균진정"; top20["G1"] = "평균좋아요/조회수"; top20["H1"] = "유튜브 컨텐츠 수"; top20["I1"] = "매출액"; top20["J1"] = "유튜브 컨텐츠 수"
   
    # ol_yu 속 데이터 추출
    brands = ol_yu["D"][1::]
    prices = ol_yu["E"][1::]
    reviews = ol_yu["G"][1::]
    points = ol_yu["H"][1::]
    treats = ol_yu["O"][1::]
    contents = ol_yu["Q"][1::]
    views = ol_yu["R"][1::]
    nices = ol_yu["S"][1::]
    repls = ol_yu["T"][1::]
    
    #trend 속 데이터 추출
    sells = trend["D"][1::]
    searchs = trend["C"][1::]
    
    # top 20 브랜드들에 대한 ol_yu속 데이터 처리
    for k in range(len(tops)):
        brand_count = 0
        count_for_nocont = 0
        count_for_noreview = 0
        sumprice = 0
        sumreview = 0
        sumpoint = 0
        sumtreat = 0
        sumnpv = 0
        sumcontents = 0
        for i in range(100):
            brand = brands[i].value
            
            if brand == tops[k]:
                
                brand_count += 1 ; count_for_nocont += 1; count_for_noreview += 1
                price = prices[i].value.replace(",","")
                sumprice += int(price)
                if reviews[i].value == -1:
                    count_for_noreview -= 1
                else:
                    sumreview += reviews[i].value
                    sumpoint += points[i].value
                    sumtreat += treats[i].value
                
                if views[i].value == 0: # 조회수가 없는, 즉 검색 결과가 없는 경우 좋아요/조회수 와 컨텐츠 수에 포함되지 않도록 설정
                    count_for_nocont -= 1
                else:
                    sumnpv += nices[i].value / views[i].value
                    sumcontents += contents[i].value
                
                
        # 구한 값들을 차례로 엑셀 칸에 집어넣음
        
        top20[f"B{k+2}"] = tops[k];
        top20[f"C{k+2}"] = sumprice / brand_count;
        
        if count_for_noreview != 0:   #zero division error 방지
            top20[f"D{k+2}"] = sumreview / count_for_noreview
            top20[f"E{k+2}"] = sumpoint / count_for_noreview
            top20[f"F{k+2}"] = sumtreat / count_for_noreview
        else:
            top20[f"D{k+2}"]=0; top20[f"E{k+2}"]=0; top20[f"F{k+2}"]=0
        
        if count_for_nocont != 0:  #zero division error 방지
            top20[f"G{k+2}"] = sumnpv / count_for_nocont
        else:
            top20[f"G{k+2}"] = 0
            
        top20[f"H{k+2}"] = sumcontents
        
        
    # top 20 브랜드들에 대한 trend속 데이터 처리
    for k in range(len(tops)):
        for i in range(len(sells)-2):
            if trend[f"B{i+2}"].value == tops[k]:
                top20[f'I{k+2}'] = sells[i].value
                top20[f'J{k+2}'] = searchs[i].value
                
    #파일 저장
    top20d.save('new_올리브영_진정_상위20.xlsx')


youtube_main():
    try :
        youtube_crawling()
        excel_add_youtube()
        make_top20()
    except Exception as e:
        print('파일 에러 발생.필요한 데이터는 아래와 같습니다.')
        print('원본_올리브영랭크100위화장품정보_상세페이지.xlsx')
        print('트렌드지수 크롤링(브랜드).xlsx')
    

#####20171228 박현진 #########################################################################################################
try:
    import plotly.express as px
    from PIL import Image
    import pandas as pd
    import matplotlib.pyplot as plt
    import seaborn as sns
    import numpy as np
    import matplotlib.font_manager as fm
    import warnings
    import plotly.graph_objects as go
    import plotly.io as po

except:
    print('!pip install plotly 를 입력하여 패키지를 설치하세요')
    print('!pip install PIL 를 입력하여 패키지를 설치하세요')
    print('!pip install warnings 를 입력하여 패키지를 설치하세요')
    
warnings.filterwarnings(action='ignore')

def visualization_by_corr(excel_file_name):
    #상관계수 표 출력
    df= pd.read_excel(excel_file_name)
    print(df.corr())
    

def visualization_by_brand(excel_file_name):
    df= pd.read_excel(excel_file_name)
    #그래프 1 : 유투브 컨텐츠 수와 올리브영 리뷰 개수, 해당 회사 매출액을 비교 
    fig = px.scatter(df, x="평균리뷰개수", y="평균좋아요/조회수",
	         size="매출액", color=df.index,
            hover_name="브랜드", log_x=True, size_max=60)
    po.write_html(fig, file='new_brand_visualization.html', auto_open=True)
    fig.show()
                  
def visualization_by_product(excel_file_name):
    df = pd.read_excel(excel_file_name)
    df=df[df['진정']>=30] #진정제품 필터링

    df['퀄리티'] = df['리뷰개수']+ df['총 댓글수']
    fig = px.scatter(
        df,
        x="퀄리티",
        y="가격",
        hover_name="상품이름",
        hover_data=['브랜드',"가격","총 댓글수", "리뷰개수",'진정']
    )
    fig.update_traces(marker_color="rgba(0,0,0,0)")

    maxDim = df[["총 댓글수",'가격']].max().idxmax()
    maxi = df[maxDim].max()
    for i, row in df.iterrows():
        rank = row['순위']
        brand = row['브랜드']
        fig.add_layout_image(
            dict(
                source=Image.open(f"원본_crawled_img/{rank}_{brand}.jpg"),
                xref="x",
                yref="y",
                xanchor="center",
                yanchor="middle",
                x=row["퀄리티"],
                y=row["가격"],
                #sizex=np.sqrt(row["진정"] / df["진정"].max()) * maxi * 0.2 + maxi * 0.05,
                #sizey=np.sqrt(row["진정"] / df["진정"].max()) * maxi * 0.2 + maxi * 0.05,
                sizex=np.sqrt(row["총 댓글수"] / df["총 댓글수"].max()) * maxi * 0.2 + maxi * 0.05,
                sizey=np.sqrt(row["총 댓글수"] / df["총 댓글수"].max()) * maxi * 0.2 + maxi * 0.05,
                sizing="contain",
                opacity=0.8,
                layer="above"
            )
        )
    fig.update_layout(height=600, width=1000, yaxis_range=[-5e3, 55e3], plot_bgcolor="#dfdfdf")
    po.write_html(fig, file='new_product_visualization.html', auto_open=True)
    fig.show()

def visualization():
    print()
    print('1. 진정제품군 중 상품별 시각화')
    print('2. 진정제품군 상위 11개 브랜드 시각화')
    
    option = int(input('원하는 시각화 옵션을 숫자로 입력하세요. : '))
    try:
        if option == 1:
            visualization_by_product("원본_올리브영랭크100위화장품정보+유튜브데이터.xlsx")
        elif option == 2:
            visualization_by_brand('원본_올리브영_진정_상위20.xlsx')
        else:
            print('1과 2중에 하나만 입력하세요.')
    except Exception as e:
        print(e)
        print('다음중 없는 파일이 존재합니다.')
        print('원본_올리브영_진정_상위20.xlsx')
        print('원본_올리브영랭크100위화장품정보+유튜브데이터.xlsx')
        print('원본_crawled_img 폴더')
        
    
if __name__=='__main__':
    from selenium.common.exceptions import TimeoutException, WebDriverException
    #1.올리브영데이터 긁기
    try: 
        olive_main()
        pass
    except WebDriverException:
        print('WebDriverException: 로딩이 느려 다시 크롤링을 시작합니다.')
        df=pd.read_excel('원본_올리브영랭크100위화장품정보.xlsx')
        print('2차 크롤링 : 올리브영랭크100위화장품정보를 읽어 상세페이지를 크롤링합니다.')
        result_df= iter_dataframe(df)
        print('올리브영 크롤링이 완료되었습니다.')
    print('WebDriverException: 로딩이 느려 다시 크롤링을 시작합니다.')
    df=pd.read_excel('원본_올리브영랭크100위화장품정보.xlsx')
    print('2차 크롤링 : 올리브영랭크100위화장품정보를 읽어 상세페이지를 크롤링합니다.')
    result_df= iter_dataframe(df)
    print('올리브영 크롤링이 완료되었습니다.')
    
    #2.아이템스카우트 긁기
    item_crawl_main()
    #3.유투브 긁기
    youtube_main()
    #상품별, 브랜드별 시각화 
    visualization()
    
